from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import tkinter as tk
import tkinter as ttk
import openpyxl
from openpyxl import Workbook
import pathlib


background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"


root = Tk()
root.title("Employee Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)


file = pathlib.Path('Employee_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    
    sheet['A1'] = "Serial Number"
    sheet['B1'] = "Name"
    sheet['C1'] = "TVSE Employee ID"
    sheet['D1'] = "TVSM EMployee ID"
    sheet['E1'] = "Team-FY'25"
    sheet['F1'] = "Designation"
    sheet['G1'] = "Role"
    sheet['H1'] = "Email ID"
    sheet['I1'] = "Team"
    sheet['J1'] = "Location"
    sheet['K1'] = "Site"
    sheet['L1'] = "DOJ"
    sheet['M1'] = "Total Expense'25"
    sheet['N1'] = "Functional Reporting TVSE"
    sheet['O1'] = "Reporting TVSE" 
    sheet['P1'] = "Department"
    sheet['Q1'] = "Blood Group"
    sheet['R1'] = "Address"
    sheet['S1'] = "Status"
    sheet['T1'] = "DOB"
    sheet['U1'] = "Gender"
    sheet['V1'] = "Shirt Size"
    sheet['W1'] = "T- Shirt Size"
    sheet['X1'] = "Joining Kit"
    sheet['Y1'] = "Qualification"
    sheet['Z1'] = "ID Card"
    sheet['AA1'] = "Canteen Recovery"
    sheet['AB1'] = "Bus Facility"
    sheet['AC1'] = "LWD"
    sheet['AD1'] = "Laptop Given"
    sheet['AE1'] = "Outside Permission"
    sheet['AF1'] = "Contact Number"
    sheet['AG1'] = "Emergency Contact Number"

    file.save('Employee_data.xlsx')

# Exit Window
def Exit():
    root.destroy()

# Show Image
def showimage():
    
    global filename
    global img

    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select Image File", filetype=(("JPG File", "*.jpg"),
                                                                               ("PNG File", "*.png"),
                                                                               ("All Files", "*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# Registration Number
def Registration_Number():
    file = openpyxl.load_workbook('Employee_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")


# Clear
def Clear():
    global img

    Name.set('')
    DOJ.set('')
    radio.set(0)  # Clear gender selection
    Location.set('')
    TVSE.set('')
    Designation.set('')
    DOB.set('')
    Contact.set('')
    Email.set('')
    Reporting.set('')
    Functional_Reporting.set('')
    Blood.set('')
    TVSM.set('')
    Address.set('')
    Status.set('')    
    Department.set('Select the Department')
    Site.set('')
    Role.set('')
    Team.set('')
    Emergency.set('')
    Team1.set('')
    Total.set('')
    Z1.set('')
    Z2.set('Select Size')
    radio1.set(1)  #Clear Joining Kit
    Z4.set('')
    radio2.set(2)  #Clear ID Card
    radio3.set(3)  #Clear Canteen
    radio4.set(4)  #Clear Bus Charge
    Z8.set('')
    Z9.set('')
    radio5.set(5)  #Clear Outside Permission

    Registration_Number()

    SaveButton.config(state='normal')
   
    img1 = PhotoImage(file="")
    lbl.config(image=img1)
    lbl.image = img1
    img = ""


# Save Button

def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Department.get()

    try:
        G1 = gender
    except NameError:
        messagebox.showerror("Error", "Select Gender")
        return

    D2 = DOJ.get()
    D1 = Date.get()
    Re1 = Location.get()
    S1 = TVSE.get()
    Bus_Facility = Designation.get()
    D3 = DOB.get()
    Food_Facility = Contact.get()
    E1 = Email.get()
    R2 = Reporting.get()
    F1 = Functional_Reporting.get()
    R3 = Role.get()
    S3 = Site.get()
    T2 = Team.get()
    T3 = Team1.get()
    T4 = Total.get()
    E2 = Emergency.get()
    B1 = Blood.get()
    T1 = TVSM.get()
    A1 = Address.get()
    S2 = Status.get()
    z1 = Z1.get()
    z2 = Z2.get()
    
    try:
        z3 = kit
    except NameError:
        messagebox.showerror("Error", "Select Joining Kit")
        return
    
    z4 = Z4.get()
    
    try:
        z5 = ID
    except NameError:
        messagebox.showerror("Error", "Select ID Detials")
        return
    
    try:
        z6 = Canteen 
    except NameError:
        messagebox.showerror("Error", "Select Canteen Details")
        return
    
    try:
        z7 = Bus 
    except NameError:
        messagebox.showerror("Error", "Select Bus Details")
        return
    
    z8 = Z8.get()
    z9 = Z9.get()
    
    try:
        z10 = Outside
    except NameError:
        messagebox.showerror("Error", "Select Outside permission")
        return
 

    if N1=="" or C1=="Select the Department" or D2=="" or Re1=="" or T2=="" or T3=="" or T4=="" or E2=="" or S1=="" or R3=="" or S3=="" or Bus_Facility=="" or D3=="" or Food_Facility=="" or E1=="" or R2=="" or F1=="" or B1=="" or T1=="" or A1=="" or S2=="" or z1=="" or z2=="" or z4==""or z8=="" or z9=="":
        messagebox.showerror("error","Few Data is Missing!")
    else:
        file=openpyxl.load_workbook('Employee_data.xlsx')
        sheet = file.active

        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=S1)
        sheet.cell(column=4,row=sheet.max_row,value=T1)
        sheet.cell(column=5,row=sheet.max_row,value=T2)
        sheet.cell(column=6,row=sheet.max_row,value=Bus_Facility)
        sheet.cell(column=7,row=sheet.max_row,value=R3)
        sheet.cell(column=8,row=sheet.max_row,value=E1)
        sheet.cell(column=9,row=sheet.max_row,value=T3)
        sheet.cell(column=10,row=sheet.max_row,value=Re1)
        sheet.cell(column=11,row=sheet.max_row,value=S3)
        sheet.cell(column=12,row=sheet.max_row,value=D2)
        sheet.cell(column=13,row=sheet.max_row,value=T4)
        sheet.cell(column=14,row=sheet.max_row,value=F1)
        sheet.cell(column=15,row=sheet.max_row,value=R2)
        sheet.cell(column=16,row=sheet.max_row,value=C1)
        sheet.cell(column=17,row=sheet.max_row,value=B1)
        sheet.cell(column=18,row=sheet.max_row,value=A1)
        sheet.cell(column=19,row=sheet.max_row,value=S2)
        sheet.cell(column=20,row=sheet.max_row,value=D3)
        sheet.cell(column=21,row=sheet.max_row,value=G1)
        sheet.cell(column=22,row=sheet.max_row,value=z1)
        sheet.cell(column=23,row=sheet.max_row,value=z2)
        sheet.cell(column=24,row=sheet.max_row,value=z3)
        sheet.cell(column=25,row=sheet.max_row,value=z4)
        sheet.cell(column=26,row=sheet.max_row,value=z5)
        sheet.cell(column=27,row=sheet.max_row,value=z6)
        sheet.cell(column=28,row=sheet.max_row,value=z7)
        sheet.cell(column=29,row=sheet.max_row,value=z8)
        sheet.cell(column=30,row=sheet.max_row,value=z9)
        sheet.cell(column=31,row=sheet.max_row,value=z10)
        sheet.cell(column=32,row=sheet.max_row,value=Food_Facility)
        sheet.cell(column=33,row=sheet.max_row,value=E2)

        file.save(r'Employee_data.xlsx')

        try:
            img.save("Emp Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile picture is not available!!!")

        messagebox.showinfo("info","Successfully data Entered!!!")
        
        Clear()
        
        Registration_Number()


#Search
def search():  
    text = Search.get()
    Clear()
    SaveButton.config(state='disable')
    
    file = openpyxl.load_workbook('Employee_data.xlsx')
    sheet = file.active
    
    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            # print(str(name)) 
            reg_position = str(name)[14:-1]
            reg_no = str(name)[15:-1]
            # print(reg_position)
            # print(reg_no)
            
    try:
        print(str(name))
    except:
        messagebox.showerror("invalid","Invalid Entries!!")
        
    x1=sheet.cell(row=int(reg_no),column=1).value
    x2=sheet.cell(row=int(reg_no),column=2).value
    x3=sheet.cell(row=int(reg_no),column=3).value
    x4=sheet.cell(row=int(reg_no),column=4).value
    x5=sheet.cell(row=int(reg_no),column=5).value
    x6=sheet.cell(row=int(reg_no),column=6).value
    x7=sheet.cell(row=int(reg_no),column=7).value
    x8=sheet.cell(row=int(reg_no),column=8).value
    x9=sheet.cell(row=int(reg_no),column=9).value
    x10=sheet.cell(row=int(reg_no),column=10).value
    x11=sheet.cell(row=int(reg_no),column=11).value
    x12=sheet.cell(row=int(reg_no),column=12).value
    x13=sheet.cell(row=int(reg_no),column=13).value
    x14=sheet.cell(row=int(reg_no),column=14).value
    x15=sheet.cell(row=int(reg_no),column=15).value
    x16=sheet.cell(row=int(reg_no),column=16).value
    x17=sheet.cell(row=int(reg_no),column=17).value
    x18=sheet.cell(row=int(reg_no),column=18).value
    x19=sheet.cell(row=int(reg_no),column=19).value
    x20=sheet.cell(row=int(reg_no),column=20).value
    x21=sheet.cell(row=int(reg_no),column=21).value
    x22=sheet.cell(row=int(reg_no),column=22).value
    x23=sheet.cell(row=int(reg_no),column=23).value
    x24=sheet.cell(row=int(reg_no),column=24).value
    x25=sheet.cell(row=int(reg_no),column=25).value
    x26=sheet.cell(row=int(reg_no),column=26).value
    x27=sheet.cell(row=int(reg_no),column=27).value
    x28=sheet.cell(row=int(reg_no),column=28).value
    x29=sheet.cell(row=int(reg_no),column=29).value
    x30=sheet.cell(row=int(reg_no),column=30).value
    x31=sheet.cell(row=int(reg_no),column=31).value
    x32=sheet.cell(row=int(reg_no),column=32).value
    x33=sheet.cell(row=int(reg_no),column=33).value

    Registration.set(x1)
    Name.set(x2)
    Department.set(x16)
    
    if x21 == 'female':
        R2.select()
    else:
        R1.select()
        
    Location.set(x10)
    TVSE.set(x3)
    Designation.set(x6)
    DOB.set(x20)
    Contact.set(x32)
    Email.set(x8)
    Functional_Reporting.set(x14)
    Reporting.set(x15)
    Blood.set(x17)
    TVSM.set(x4)
    Address.set(x18)
    Status.set(x19)    
    DOJ.set(x12)
    Z1.set(x22)
    Z2.set(x23)
    
    if x24 == "No":
        Z32.select()
    else:
        Z31.select()
        
    Z4.set(x25)
    
    if x26 == "No":
        Z52.select()
    else:
        Z51.select()
        
    if x27 == "No":
        Z62.select()
    else:
        Z61.select()
        
    if x28 == "No":
        Z72.select()
    else:
        Z71.select()
        
    Z8.set(x29)
    Z9.set(x30)
    
    if x31 == "No":
        Z102.select()
    else:
        Z101.select()
    
    Role.set(x7)
    Site.set(x11)
    Team.set(x5)
    Emergency.set(x33)
    Team1.set(x9)
    Total.set(x13)
    
    img = (Image.open('Emp Images/'+str(x1)+".jpg"))
    resized_image = img.resize((200,200))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    
#Update
def update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Department.get()
    
    selection()
    G1=gender
    
    D2 = DOJ.get()
    D1 = Date.get()
    Re1 = Location.get()
    S1 = TVSE.get()
    Bus_Facility = Designation.get()
    D3 = DOB.get()
    Food_Facility = Contact.get()
    E1 = Email.get()
    R2 = Reporting.get()
    F1 = Functional_Reporting.get()
    B1 = Blood.get()
    T1 = TVSM.get()
    A1 = Address.get()
    S2 = Status.get()
    R3 = Role.get()
    S3 = Site.get()
    T2 = Team.get()
    E2 = Emergency.get()
    T3 = Team1.get()
    T4 = Total.get()
    z1 = Z1.get()
    z2 = Z2.get()
    
    selection1()
    z3 = kit
    
    z4 = Z4.get()
    
    selection2()
    z5 = ID
    
    selection3()
    z6 = Canteen
    
    selection4()
    z7 = Bus
    
    z8 = Z8.get()
    z9 = Z9.get()
    
    selection5()
    z10 = Outside 
    
    file = openpyxl.load_workbook("Employee_data.xlsx")
    sheet = file.active
    
    for row in sheet.rows:
        if row[0].value ==R1:
            name=row[0]
            print(str(name))
            reg_no_pos = str(name)[14:-1]
            reg_no = str(name)[15:-1]
            
            print(reg_no)
            
    sheet.cell(column=1,row=int(reg_no),value=R1)
    sheet.cell(column=2,row=int(reg_no),value=N1)
    sheet.cell(column=3,row=int(reg_no),value=S1)
    sheet.cell(column=4,row=int(reg_no),value=T1)
    sheet.cell(column=5,row=int(reg_no),value=T2)
    sheet.cell(column=6,row=int(reg_no),value=Bus_Facility)
    sheet.cell(column=7,row=int(reg_no),value=R3)
    sheet.cell(column=8,row=int(reg_no),value=E1)
    sheet.cell(column=9,row=int(reg_no),value=T3)
    sheet.cell(column=10,row=int(reg_no),value=Re1)
    sheet.cell(column=11,row=int(reg_no),value=S3)
    sheet.cell(column=12,row=int(reg_no),value=D2)
    sheet.cell(column=13,row=int(reg_no),value=T4)
    sheet.cell(column=14,row=int(reg_no),value=F1)
    sheet.cell(column=15,row=int(reg_no),value=R2)
    sheet.cell(column=16,row=int(reg_no),value=C1)
    sheet.cell(column=17,row=int(reg_no),value=B1)
    sheet.cell(column=18,row=int(reg_no),value=A1)
    sheet.cell(column=19,row=int(reg_no),value=S2)
    sheet.cell(column=20,row=int(reg_no),value=D3)
    sheet.cell(column=21,row=int(reg_no),value=G1)
    sheet.cell(column=22,row=int(reg_no),value=z1)
    sheet.cell(column=23,row=int(reg_no),value=z2)
    sheet.cell(column=24,row=int(reg_no),value=z3)
    sheet.cell(column=25,row=int(reg_no),value=z4)
    sheet.cell(column=26,row=int(reg_no),value=z5)
    sheet.cell(column=27,row=int(reg_no),value=z6)
    sheet.cell(column=28,row=int(reg_no),value=z7)
    sheet.cell(column=29,row=int(reg_no),value=z8)
    sheet.cell(column=30,row=int(reg_no),value=z9)
    sheet.cell(column=31,row=int(reg_no),value=z10)
    sheet.cell(column=32,row=int(reg_no),value=Food_Facility)
    sheet.cell(column=33,row=int(reg_no),value=E2)
    
    file.save(r'Employee_data.xlsx')
    
    try:
        img.save('Emp Images/'+str(R1)+".jpg")
    except:
        pass
    
    messagebox.showinfo("Update","Updated Successfully!!")
    Clear()
    
        
# Gender Selection
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"
        
#Joining Kit
def selection1():
    global kit 
    value = radio1.get()
    if value ==1:
        kit = "yes"
    else:
        kit = "No"
        
#ID Card
def selection2():
    global ID 
    value = radio2.get()
    if value ==1:
        ID = "yes"
    else:
        ID = "No"
        
#Canteen 
def selection3():
    global Canteen 
    value = radio3.get()
    if value ==1:
        Canteen = "yes"
    else:
        Canteen = "No"
        
#Bus Charge
def selection4():
    global Bus 
    value = radio4.get()
    if value ==1:
        Bus = "yes"
    else:
        Bus = "No"
        
#Outside Permission
def selection5():
    global Outside  
    value = radio5.get()
    if value ==1:
        Outside = "yes"
    else:
        Outside = "No"
        
# Global Excel file path
excel_file = "Employee_data.xlsx"
   
#Export Data 
def export_selected_fields_ui(master, excel_path):
    def export_selected():
        selected_cols = [col for col, var in zip(columns, vars_list) if var.get() == 1]
        if not selected_cols:
            messagebox.showwarning("No Selection", "Please select at least one field to export.")
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]
            col_indexes = [headers.index(col) for col in selected_cols]

            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Exported Data"

            for idx, col in enumerate(selected_cols, start=1):
                new_ws.cell(row=1, column=idx, value=col)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                for col_num, col_index in enumerate(col_indexes, start=1):
                    new_ws.cell(row=row_idx, column=col_num, value=row[col_index])

            export_path = os.path.join(os.path.dirname(excel_path), "Exported_Employees_Details.xlsx")
            new_wb.save(export_path)
            messagebox.showinfo("Success", f"Exported to {export_path}")
            export_window.destroy()

        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    # Read columns from Excel
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        columns = [cell.value for cell in sheet[1]]
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read Excel file: {e}")
        return

    # Create compact popup window
    export_window = Toplevel(master)
    export_window.title("Select Fields to Export")
    export_window.geometry("580x450")
    export_window.configure(bg="#f8f8f8")

    # Canvas + Scrollbar
    canvas = Canvas(export_window, borderwidth=0, bg="#f8f8f8")
    scrollbar = Scrollbar(export_window, orient="vertical", command=canvas.yview)
    scrollable_frame = Frame(canvas, bg="#f8f8f8")

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Horizontal grouped boxes in a compact layout
    container_frame = Frame(scrollable_frame, bg="#f8f8f8")
    container_frame.pack(padx=5, pady=5)

    vars_list = []
    chunk_size = 11
    for i in range(0, len(columns), chunk_size):
        group_frame = LabelFrame(container_frame,
                                 text=f"Fields {i + 1}-{min(i + chunk_size, len(columns))}",
                                 padx=5, pady=5,
                                 bg="#eaeaea",
                                 font=("Segoe UI", 9, "bold"))
        group_frame.grid(row=0, column=i // chunk_size, padx=5, pady=5, sticky="n")

        for idx, col in enumerate(columns[i:i + chunk_size]):
            var = IntVar()
            chk = Checkbutton(group_frame, text=col, variable=var, bg="#eaeaea", font=("Segoe UI", 9))
            chk.pack(anchor="w", pady=1)
            vars_list.append(var)

    # Export Button (compact spacing)
    button_frame = Frame(scrollable_frame, bg="#f8f8f8")
    button_frame.pack(pady=10)

    export_btn = ttk.Button(button_frame, text="Export Selected Fields", command=export_selected, bg='black', fg='white',font=("Segoe UI", 9, "bold"))
    export_btn.pack()

    # Button style
    style = ttk.Style()
    style.configure("TButton", font=("Segoe UI", 9, "bold"), padding=2)

# Top Frames
Label(root, text="Email: Sk.rajeshkannan@tvsmotor.com", width=10, height=2, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="Employee Registration : ", width=10, height=2, bg="#c36464", font="arial 20 bold").pack(side=TOP, fill=X)

#Search box to Update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=925,y=55)
imageicon3=PhotoImage(file="")
Srch = Button(root,text="Search",compound=LEFT,image=imageicon3,width=70,bg='#68ddfa',font="arial 13 bold",command=search)
Srch.place(x=1180,y=60)

# imageicon4 = PhotoImage(file="")
# Update_Button = Button(root,image=imageicon4,bg="#c36464")
# Update_Button.place(x=1200,y=64)

imageicon4 = PhotoImage(file="")
Updt = Button(root,text="Update",compound=LEFT,image=imageicon4,width=70,bg='#68ddfa',font="arial 13 bold",command=update)
Updt.place(x=1300,y=60)

# Registration and Date

Label(root, text="Registration Number:", font="arial 13", fg=framebg, bg=background).place(x=2, y=130)
Label(root, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=130)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=130)

Registration_Number()


today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=130)

Date.set(d1)


# Employee Details
obj = LabelFrame(root, text="Work Profile", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=310, relief=GROOVE)
obj.place(x=30, y=180)


Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=25)
Label(obj, text="DOJ:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=65)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=105)
Label(obj, text="Funct Reporting:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=145)
Label(obj, text="Role:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=185)
Label(obj, text="Total Exp:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=225)

Label(obj, text="Department:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=25)
Label(obj, text="Location:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=65)
Label(obj, text="TVSE Emp No. :", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=105)
Label(obj, text="Reporting-TVSE:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=145)
Label(obj, text="Site:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=185)
Label(obj, text="Team:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=225)
 

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=25)

DOJ = StringVar()
doj_entry = Entry(obj, textvariable=DOJ, width=20, font="arial 10")
doj_entry.place(x=160, y=65)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=160, y=105)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=250, y=105)

Functional_Reporting = StringVar()
Reporting_TVSM_entry = Entry(obj, textvariable=Functional_Reporting, width=20, font="arial 10")
Reporting_TVSM_entry.place(x=160, y=150)

Role = StringVar()
Role_entry = Entry(obj, textvariable=Role, width=20, font="arial 10")
Role_entry.place(x=160, y=190)

Total = StringVar()
Total_entry = Entry(obj, textvariable=Total, width=20, font="arial 10")
Total_entry.place(x=160, y=230)

Location = StringVar()
Location_entry = Entry(obj, textvariable=Location, width=20, font="arial 10")
Location_entry.place(x=630, y=65)

TVSE = StringVar()
Site_entry = Entry(obj, textvariable=TVSE, width=20, font="arial 10")
Site_entry.place(x=630, y=105)

Department = Combobox(obj, values=['D&AI', 'IT', 'SALES', 'CP', 'PED', 'OPNS', 'R&D', 'SERVICE', 'IB', 'NPD', 'QAD',
                                   'HRD', 'CMD', 'FINANCE', 'COMMOBILITY', 'SUSTAINABILITY', 'CU', 'SMD', 'CIVIL',
                                   'CKD'], font="Roboto 10", width=17, state="r")

Department.place(x=630, y=30)
Department.set("Select the Department")

Reporting = StringVar()
Reporting_TVSE_entry = Entry(obj, textvariable=Reporting, width=20, font="arial 10")
Reporting_TVSE_entry.place(x=630, y=150)

Site = StringVar()
Site_entry = Entry(obj, textvariable=Site, width=20, font="arial 10")
Site_entry.place(x=630, y=190)

Team1 = StringVar()
Team1_entry = Entry(obj, textvariable=Team1, width=20, font="arial 10")
Team1_entry.place(x=630, y=230)

# Other Details
obj2 = LabelFrame(root, text="Employee Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=280, relief=GROOVE)
obj2.place(x=30, y=500)

Label(obj2, text="Designation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=10)
Label(obj2, text="Contact Number:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=60)
Label(obj2, text="TVSM Emp No. :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=110)
Label(obj2, text="Blood Group :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=160)
Label(obj2, text="Team-FY'25 : ", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=210)

Designation = StringVar()
Bus_entry = Entry(obj2, textvariable=Designation, width=20, font="arial 10")
Bus_entry.place(x=160, y=15)

Contact = StringVar()
Food_entry = Entry(obj2, textvariable=Contact, width=20, font="arial 10")
Food_entry.place(x=160, y=65)

TVSM = StringVar()
TVSM_entry = Entry(obj2, textvariable=TVSM, width=20, font="arial 10")
TVSM_entry.place(x=160, y=115)
 
Blood = StringVar()
Blood_entry = Entry(obj2, textvariable=Blood, width=20, font="arial 10")
Blood_entry.place(x=160, y=165)

Team = StringVar()
Team_entry = Entry(obj2, textvariable=Team, width=20, font="arial 10")
Team_entry.place(x=160, y=215)

Label(obj2, text="DOB:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=10)
Label(obj2, text="Email Address:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=60)
Label(obj2, text="Address:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=110) 
Label(obj2, text="Status:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=160)
Label(obj2, text="Emergency No. :", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=210)

DOB = StringVar()
DOB_entry = Entry(obj2, textvariable=DOB, width=20, font="arial 10")
DOB_entry.place(x=630, y=15) 

Email = StringVar()
Email_entry = Entry(obj2, textvariable=Email, width=20, font="arial 10")
Email_entry.place(x=630, y=65)

Address = StringVar()
Address_entry = Entry(obj2, textvariable=Address, width=20, font="arial 10")
Address_entry.place(x=630, y=115) 

Status = StringVar()
Status_entry = Entry(obj2, textvariable=Status, width=20, font="arial 10")
Status_entry.place(x=630, y=165)

Emergency = StringVar()
Emergency_entry = Entry(obj2, textvariable=Emergency, width=20, font="arial 10")
Emergency_entry.place(x=630, y=215)

# Other Details
obj3 = LabelFrame(root, text="Other Details", font=20, bd=2, width=380, bg=framebg, fg=framefg, height=600, relief=GROOVE)
obj3.place(x=940, y=180)

Label(obj3, text="Shirt Size:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=30)              #Z1
Label(obj3, text="T - Shirt Size:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=80)          #Z2
Label(obj3, text="Joining Kit:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=130)            #Z3 
Label(obj3, text="Qualification :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=180)         #Z4
Label(obj3, text="ID Card:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=230)                #Z5
Label(obj3, text="Canteen Recovery:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=280)       #Z6 
Label(obj3, text="Bus Facility :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=330)          #Z7
Label(obj3, text="LWD :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=380)                   #Z8
Label(obj3, text="Laptop Number :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=430)         #Z9
Label(obj3, text="Outside Permission :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=480)    #Z10

Z1 = StringVar()
DOB_entry = Entry(obj3, textvariable=Z1, width=20, font="arial 10")
DOB_entry.place(x=200, y=35)

# Z2 = StringVar()
# DOB_entry = Entry(obj3, textvariable=Z2, width=20, font="arial 10")
# DOB_entry.place(x=200, y=85)

Z2 = Combobox(obj3, values=['S','M','L','XL','XXL','XXXL'], font="Roboto 10", width=17, state="r")

Z2.place(x=200, y=85)
Z2.set("Select Size")

# Z3 = StringVar()
# DOB_entry = Entry(obj3, textvariable=Z3, width=20, font="arial 10")
# DOB_entry.place(x=200, y=135)

radio1 = IntVar()
Z31 = Radiobutton(obj3, text="Yes", variable=radio1, value=1, bg=framebg, fg=framefg, command=selection1)
Z31.place(x=200, y=130)

Z32 = Radiobutton(obj3, text="No", variable=radio1, value=2, bg=framebg, fg=framefg, command=selection1)
Z32.place(x=300, y=130)

Z4 = StringVar()
DOB_entry = Entry(obj3, textvariable=Z4, width=20, font="arial 10")
DOB_entry.place(x=200, y=185)

# Z5 = StringVar()
# DOB_entry = Entry(obj3, textvariable=Z5, width=20, font="arial 10")
# DOB_entry.place(x=200, y=235)

radio2 = IntVar()
Z51 = Radiobutton(obj3, text="Yes", variable=radio2, value=1, bg=framebg, fg=framefg, command=selection2)
Z51.place(x=200, y=235)

Z52 = Radiobutton(obj3, text="No", variable=radio2, value=2, bg=framebg, fg=framefg, command=selection2)
Z52.place(x=300, y=235)

# Z6 = StringVar()
# DOB_entry = Entry(obj3, textvariable=Z6, width=20, font="arial 10")
# DOB_entry.place(x=200, y=285)

radio3 = IntVar()
Z61 = Radiobutton(obj3, text="Yes", variable=radio3, value=1, bg=framebg, fg=framefg, command=selection3)
Z61.place(x=200, y=285)

Z62 = Radiobutton(obj3, text="No", variable=radio3, value=2, bg=framebg, fg=framefg, command=selection3)
Z62.place(x=300, y=285)

# Z7 = StringVar()
# DOB_entry = Entry(obj3, textvariable=Z7, width=20, font="arial 10")
# DOB_entry.place(x=200, y=335)

radio4 = IntVar()
Z71 = Radiobutton(obj3, text="Yes", variable=radio4, value=1, bg=framebg, fg=framefg, command=selection4)
Z71.place(x=200, y=335)

Z72 = Radiobutton(obj3, text="No", variable=radio4, value=2, bg=framebg, fg=framefg, command=selection4)
Z72.place(x=300, y=335)

Z8 = StringVar()
DOB_entry = Entry(obj3, textvariable=Z8, width=20, font="arial 10")
DOB_entry.place(x=200, y=385)

Z9 = StringVar()
DOB_entry = Entry(obj3, textvariable=Z9, width=20, font="arial 10")
DOB_entry.place(x=200, y=435)

# Z10 = StringVar()
# DOB_entry = Entry(obj3, textvariable=Z10, width=20, font="arial 10")
# DOB_entry.place(x=200, y=485)

radio5 = IntVar()
Z101 = Radiobutton(obj3, text="Yes", variable=radio5, value=1, bg=framebg, fg=framefg, command=selection5)
Z101.place(x=200, y=485)

Z102 = Radiobutton(obj3, text="No", variable=radio5, value=2, bg=framebg, fg=framefg, command=selection5)
Z102.place(x=300, y=485)

# Image Frame
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1330, y=200)

lbl = Label(f, bg="black")
lbl.place(x=0, y=0)

# Buttons
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1330, y=440)
SaveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
SaveButton.place(x=1330, y=520)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1330 , y=600)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1330, y=680)
tk.Button(root, text="Export Selected Fields",width=19, height=1, font="arial 12 bold", bg="lightgreen" ,command=lambda: export_selected_fields_ui(root, excel_file)).place(x=1300, y=130)

root.mainloop()
